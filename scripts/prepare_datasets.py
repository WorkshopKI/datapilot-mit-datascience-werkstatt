#!/usr/bin/env python3
"""
Lädt Berlin Open Data herunter und konvertiert in DataPilot-kompatible CSVs.
Einmalig ausführen, Ergebnis in public/data/ einchecken.

Voraussetzungen: pip install requests openpyxl
"""

import csv
import io
import os
from collections import defaultdict
from datetime import datetime

try:
    import requests
    import openpyxl
except ImportError:
    print("pip install requests openpyxl")
    exit(1)

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'public', 'data')
os.makedirs(OUTPUT_DIR, exist_ok=True)

LOR_BEZIRK = {
    '01': 'Mitte', '02': 'Friedrichshain-Kreuzberg', '03': 'Pankow',
    '04': 'Charlottenburg-Wilmersdorf', '05': 'Spandau', '06': 'Steglitz-Zehlendorf',
    '07': 'Tempelhof-Schoeneberg', '08': 'Neukoelln', '09': 'Treptow-Koepenick',
    '10': 'Marzahn-Hellersdorf', '11': 'Lichtenberg', '12': 'Reinickendorf',
}
WOCHENTAGE = ['Montag', 'Dienstag', 'Mittwoch', 'Donnerstag', 'Freitag', 'Samstag', 'Sonntag']


def prepare_kfz():
    """Kfz-Diebstahl: Pipe-CSV → Komma-CSV mit Feature Engineering."""
    print("Downloading Kfz-Diebstahl...")
    r = requests.get('https://www.polizei-berlin.eu/Kfzdiebstahl/Kfzdiebstahl.csv')
    r.encoding = 'latin-1'
    lines = r.text.strip().split('\n')

    header = 'TATZEIT_ANFANG_DATUM,TATZEIT_ANFANG_STUNDE,WOCHENTAG,MONAT,LOR,BEZIRK,SCHADENSHOEHE,VERSUCH,DELIKT,EINDRINGEN_IN_KFZ,ERLANGTES_GUT'
    rows = [header]
    for line in lines[1:]:
        line = line.strip('\r')
        if not line:
            continue
        cols = line.split('|')
        if len(cols) < 11:
            continue
        try:
            parts = cols[1].split('.')
            dt = datetime(int(parts[2]), int(parts[1]), int(parts[0]))
            wt = WOCHENTAGE[dt.weekday()]
            monat = int(parts[1])
        except (ValueError, IndexError):
            continue
        lor = cols[5]
        bezirk = LOR_BEZIRK.get(lor[:2], 'Unbekannt')
        esc = lambda s: f'"{s}"' if ',' in s else s
        rows.append(','.join([
            cols[1], cols[2], wt, str(monat),
            lor, bezirk, cols[6], cols[7],
            esc(cols[8]), esc(cols[9]), esc(cols[10].strip('\r'))
        ]))
    path = os.path.join(OUTPUT_DIR, 'berlin-kfz-diebstahl.csv')
    with open(path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(rows))
    print(f"  → {len(rows)-1} rows → {path}")


def prepare_abwasser():
    """Abwassermonitoring: Long→Wide Format, deutsche Dezimalkommas → Punkt."""
    print("Downloading Abwasser-Viruslast...")
    r = requests.get('https://data.lageso.de/infektionsschutz/opendata/abwassermonitoring/BEWAC_abwassermonitoring_berlin.csv')
    r.encoding = 'utf-8-sig'
    reader = csv.DictReader(io.StringIO(r.text), delimiter=';')

    pivot = defaultdict(dict)
    for row in reader:
        key = (row['Datum'], row['UWW_Name'])
        dec = lambda s: s.replace(',', '.')
        pivot[key]['Datum'] = row['Datum']
        pivot[key]['Klaerwerk'] = row['UWW_Name']
        pivot[key]['Durchfluss'] = dec(row['Durchfluss'])
        pivot[key]['Temperatur'] = dec(row['Abwasser_Temperatur'])
        pivot[key]['pH'] = dec(row['Abwasser_pH'])
        erreger = row['Erreger'].replace(' ', '_')
        pivot[key][f'Viruslast_{erreger}'] = dec(row['Messwert'])

    header = 'Datum,Klaerwerk,Durchfluss,Temperatur,pH,Monat,Kalenderwoche,Viruslast_Influenza_A,Viruslast_Influenza_B,Viruslast_RSV'
    rows_out = [header]
    for data in sorted(pivot.values(), key=lambda d: (d['Datum'], d['Klaerwerk'])):
        dt = datetime.strptime(data['Datum'], '%Y-%m-%d')
        monat = dt.month
        kw = dt.isocalendar()[1]
        rows_out.append(','.join([
            data['Datum'], data['Klaerwerk'], data.get('Durchfluss', ''),
            data.get('Temperatur', ''), data.get('pH', ''),
            str(monat), str(kw),
            data.get('Viruslast_Influenza_A', ''), data.get('Viruslast_Influenza_B', ''),
            data.get('Viruslast_RSV', '')
        ]))
    path = os.path.join(OUTPUT_DIR, 'berlin-abwasser-viruslast.csv')
    with open(path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(rows_out))
    print(f"  → {len(rows_out)-1} rows → {path}")


def prepare_kriminalitaetsatlas():
    """Kriminalitätsatlas: XLSX → CSV, nur Bezirksregionen (nicht Summen)."""
    print("Downloading Kriminalitätsatlas...")
    url = 'https://www.kriminalitaetsatlas.berlin.de/K-Atlas/bezirke/Fallzahlen&HZ 2015-2024.xlsx'
    r = requests.get(url)
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True)
    ws = wb['Fallzahlen_2024']

    all_rows = list(ws.iter_rows(values_only=True))
    clean_header = [
        'LOR_Schluessel', 'Bezeichnung', 'Straftaten_insgesamt', 'Raub',
        'Strassenraub', 'Koerperverletzung_insgesamt', 'Gefaehrliche_KV',
        'Noetigung_Bedrohung', 'Wohnungseinbruch', 'Diebstahl_insgesamt',
        'Kfz_Diebstahl', 'Diebstahl_aus_Kfz', 'Fahrrad_Diebstahl',
        'Taschendiebstahl', 'Brandstiftung', 'Sachbeschaedigung_Kfz',
        'Sachbeschaedigung_sonstige', 'Rauschgiftkriminalitaet', 'Betrug'
    ]
    rows_out = [','.join(clean_header)]
    for row in all_rows[5:]:
        if not row[0]:
            continue
        lor = str(row[0]).strip()
        # Skip Bezirks-Summen (4-digit codes ending 0000) and Berlin total
        if lor.endswith('0000') or (lor.endswith('00') and len(lor) <= 4):
            continue
        # Only keep 6-digit Bezirksregionen
        if len(lor) != 6:
            continue
        vals = []
        for v in row[:len(clean_header)]:
            if v is None:
                vals.append('')
            else:
                s = str(v)
                vals.append(f'"{s}"' if ',' in s else s)
        rows_out.append(','.join(vals))
    path = os.path.join(OUTPUT_DIR, 'berlin-kriminalitaetsatlas-2024.csv')
    with open(path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(rows_out))
    print(f"  → {len(rows_out)-1} rows → {path}")


def prepare_radzaehldaten():
    """Radzähldaten: XLSX → CSV, 1 Zählstelle (Jannowitzbrücke), langes Format."""
    print("Downloading Radzähldaten (16 MB XLSX)...")
    url = 'https://www.berlin.de/sen/uvk/_assets/verkehr/verkehrsplanung/radverkehr/weitere-radinfrastruktur/zaehlstellen-und-fahrradbarometer/gesamtdatei-stundenwerte.xlsx'
    r = requests.get(url)
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True)
    ws = wb['Jahresdatei 2024']

    all_rows = list(ws.iter_rows(values_only=True))
    header_row = all_rows[0]
    # Find Jannowitzbrücke Nord column
    jan_col = None
    for idx, cell in enumerate(header_row):
        if cell and 'JAN-N' in str(cell):
            jan_col = idx
            break
    if jan_col is None:
        print("  Jannowitzbrücke nicht gefunden, nehme Spalte 2")
        jan_col = 2

    header = 'Datum,Stunde,Zaehlstelle,Anzahl,Wochentag,Monat,Ist_Wochenende'
    rows_out = [header]
    for row in all_rows[1:]:
        dt = row[0]
        if dt is None or not hasattr(dt, 'strftime'):
            continue
        count = row[jan_col]
        if count is None:
            continue
        wt = WOCHENTAGE[dt.weekday()]
        is_we = 'Ja' if dt.weekday() >= 5 else 'Nein'
        rows_out.append(','.join([
            dt.strftime('%Y-%m-%d'), str(dt.hour),
            'Jannowitzbruecke_Nord', str(int(count)),
            wt, str(dt.month), is_we
        ]))
    path = os.path.join(OUTPUT_DIR, 'berlin-radzaehldaten-2024.csv')
    with open(path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(rows_out))
    print(f"  → {len(rows_out)-1} rows → {path}")


if __name__ == '__main__':
    prepare_kfz()
    prepare_abwasser()
    prepare_kriminalitaetsatlas()
    prepare_radzaehldaten()
    print("\nAlle Datensätze aufbereitet!")
